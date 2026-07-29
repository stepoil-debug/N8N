from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from . import audit_service

_ORIGINAL_NORMALIZE = audit_service.normalize_analysis
_ORIGINAL_WRITE_EXCEL = audit_service._write_excel


def _objects(value: Any) -> list[dict[str, Any]]:
    return [item for item in (value or []) if isinstance(item, dict)]


def _drawing_metrics(items: list[dict[str, Any]]) -> dict[str, int]:
    issues = [issue for drawing in items for issue in _objects(drawing.get("issues"))]
    return {
        "drawings_analyzed": sum(1 for item in items if item.get("status") == "analyzed"),
        "drawing_pages_analyzed": sum(int(item.get("pages_analyzed") or 0) for item in items),
        "drawing_issues_total": len(issues),
        "drawing_candidate_findings": sum(1 for issue in issues if issue.get("status") == "candidate_finding"),
        "drawing_not_verifiable": sum(1 for issue in issues if issue.get("status") != "candidate_finding"),
        "drawing_blocking_candidates": sum(1 for issue in issues if issue.get("blocking")),
    }


def normalize_with_drawings(
    analysis: dict[str, Any],
    opportunity: dict[str, Any],
    package: dict[str, Any],
) -> dict[str, Any]:
    data = _ORIGINAL_NORMALIZE(analysis, opportunity, package)
    drawings = _objects(analysis.get("drawing_analysis"))
    data["drawing_analysis"] = drawings
    data["summary"].update(_drawing_metrics(drawings))
    data["batch_info"] = analysis.get("batch_info") if isinstance(analysis.get("batch_info"), dict) else {}
    return data


def _add_drawing_sheet(path: Path, data: dict[str, Any]) -> None:
    workbook = load_workbook(path)
    if "Desenhos" in workbook.sheetnames:
        del workbook["Desenhos"]
    sheet = workbook.create_sheet("Desenhos")
    headers = [
        "Documento",
        "Origem",
        "Status",
        "Página",
        "Região",
        "Regra",
        "Severidade",
        "Título",
        "Evidência visual",
        "Contradição",
        "Correção necessária",
        "Confiança",
        "Bloqueante",
        "Validação humana",
    ]
    sheet.append(headers)
    navy = PatternFill("solid", fgColor="0B2D4D")
    for cell in sheet[1]:
        cell.fill = navy
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for drawing in _objects(data.get("drawing_analysis")):
        issues = _objects(drawing.get("issues"))
        if not issues:
            sheet.append(
                [
                    drawing.get("source_document"),
                    drawing.get("source_owner"),
                    drawing.get("status"),
                    "",
                    "",
                    "",
                    "",
                    "Sem achado visual confirmado",
                    "; ".join(str(item) for item in (drawing.get("warnings") or [])),
                    "",
                    "",
                    "",
                    "NÃO",
                    "SIM" if drawing.get("status") != "analyzed" else "NÃO",
                ]
            )
        for issue in issues:
            sheet.append(
                [
                    issue.get("source_document") or drawing.get("source_document"),
                    issue.get("source_owner") or drawing.get("source_owner"),
                    issue.get("status"),
                    issue.get("page"),
                    issue.get("region"),
                    issue.get("rule_id"),
                    issue.get("severity"),
                    issue.get("title"),
                    issue.get("evidence"),
                    issue.get("contradiction"),
                    issue.get("required_correction"),
                    issue.get("confidence"),
                    "SIM" if issue.get("blocking") else "NÃO",
                    "SIM" if issue.get("requires_human_review") else "NÃO",
                ]
            )

    widths = [42, 14, 20, 10, 18, 12, 16, 42, 65, 55, 65, 12, 14, 18]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    summary = workbook["Resumo"] if "Resumo" in workbook.sheetnames else None
    if summary is not None:
        metrics = _drawing_metrics(_objects(data.get("drawing_analysis")))
        summary.append(["Desenhos analisados", metrics["drawings_analyzed"]])
        summary.append(["Páginas de desenho analisadas", metrics["drawing_pages_analyzed"]])
        summary.append(["Achados visuais candidatos", metrics["drawing_candidate_findings"]])
        summary.append(["Pontos visuais não verificáveis", metrics["drawing_not_verifiable"]])

    workbook.save(path)


def write_excel_with_drawings(path: Path, data: dict[str, Any]) -> None:
    _ORIGINAL_WRITE_EXCEL(path, data)
    _add_drawing_sheet(path, data)


audit_service.normalize_analysis = normalize_with_drawings
audit_service._write_excel = write_excel_with_drawings
