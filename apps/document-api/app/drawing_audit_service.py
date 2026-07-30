from __future__ import annotations

import io
import json
import os
import re
import zipfile
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from xml.sax.saxutils import escape

from fastapi import Depends, File, Form, HTTPException, UploadFile
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .ai_batch_service import chunks, clean, entry_text, model, objects
from .client_drawing_vision import analyze_drawing_with_client_context
from .client_knowledge import (
    _active_rules,
    _context_alerts,
    _detect_area,
    _detected_classes,
    _detected_services,
    _haystack,
    load_client_profiles,
    resolve_client_knowledge,
)
from .drawing_vision import drawing_findings, prepare_drawing_visuals
from .main import app, artifact, auth, safe, workspace
from .package_service import analyze_package

MAX_DRAWINGS = int(os.getenv("DRAWING_AUDIT_MAX_FILES", "24"))
MAX_CONDITION_FILES = int(os.getenv("DRAWING_AUDIT_MAX_CONDITION_FILES", "10"))
MAX_CONDITION_CHUNKS = int(os.getenv("DRAWING_AUDIT_MAX_CONDITION_CHUNKS", "10"))
DRAWING_EXTENSIONS = {".pdf", ".png", ".jpg", ".jpeg", ".webp", ".tif", ".tiff"}
CONDITION_EXTENSIONS = {".pdf", ".docx", ".xlsx", ".xlsm", ".csv", ".txt", ".md", ".json"}
CONDITION_TERMS = (
    "condicional", "conditional", "requisito", "requirement", "specification", "especificacao",
    "especificação", "standard", "norma", "piping material class", "material classes", "pmc",
    "data sheet", "datasheet", "project requirement", "client rule", "criterio", "critério",
)


def _fold(value: Any) -> str:
    text = str(value or "").casefold()
    return text.replace("ç", "c").replace("ã", "a").replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")


def _condition_candidate(entry: dict[str, Any]) -> bool:
    if str(entry.get("extension") or "").casefold() not in CONDITION_EXTENSIONS:
        return False
    path_text = _fold(entry.get("path"))
    preview_text = _fold(entry_text(entry)[:5000])
    if any(term in path_text for term in CONDITION_TERMS):
        return True
    signatures = (
        "piping material classes specification",
        "revision status / summary of changes",
        "normative references",
        "material substitutions",
        "project requirements",
        "contractor data requirements",
        "appendix b. notes",
    )
    return any(signature in preview_text for signature in signatures)


def _classify_drawing_package(package: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    entries = [entry for entry in package.get("entries") or [] if isinstance(entry, dict)]
    conditionals = [entry for entry in entries if _condition_candidate(entry)]
    conditional_paths = {str(entry.get("path")) for entry in conditionals}
    drawings = [
        entry for entry in entries
        if str(entry.get("extension") or "").casefold() in DRAWING_EXTENSIONS
        and str(entry.get("path")) not in conditional_paths
    ]
    other = [entry for entry in entries if entry not in drawings and entry not in conditionals]
    for entry in drawings:
        entry["document_type"] = "drawing"
        entry["source_owner"] = "client"
        entry["group"] = "drawings"
    for entry in conditionals:
        entry["document_type"] = "client_condition"
        entry["source_owner"] = "client"
        entry["group"] = "client_conditions"
    return drawings, conditionals, other


def _prepare_missing_visuals(zip_bytes: bytes, drawings: list[dict[str, Any]], opportunity_id: str) -> list[str]:
    warnings: list[str] = []
    selected = drawings[:MAX_DRAWINGS]
    if len(drawings) > MAX_DRAWINGS:
        warnings.append(f"O ZIP contém {len(drawings)} desenhos; esta execução foi limitada aos primeiros {MAX_DRAWINGS}.")
    by_path = {str(entry.get("path")): entry for entry in selected}
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as archive:
        names = {name.replace("\\", "/").lstrip("/"): name for name in archive.namelist()}
        for path, entry in by_path.items():
            if (entry.get("drawing_visuals") or {}).get("status") == "prepared":
                continue
            archive_name = names.get(path)
            if not archive_name:
                warnings.append(f"Arquivo não localizado no ZIP para renderização: {path}")
                continue
            try:
                raw = archive.read(archive_name)
                entry["drawing_visuals"] = prepare_drawing_visuals(str(entry.get("filename") or path), raw, opportunity_id)
            except Exception as exc:  # noqa: BLE001
                warnings.append(f"Falha ao preparar {path}: {exc}")
    return warnings


def _agent_metadata(agents: list[Any]) -> dict[str, str]:
    result: dict[str, str] = {}
    for item in agents:
        value = str(item or "")
        for prefix, key in (("client-profile:", "profile_id"), ("area:", "area"), ("project:", "project")):
            if value.startswith(prefix):
                result[key] = value.split(":", 1)[1].strip()
    return result


def _profile_by_id(profile_id: str) -> dict[str, Any] | None:
    return next((profile for profile in load_client_profiles() if profile.get("profile_id") == profile_id), None)


def _condition_selection(text: str) -> list[str]:
    parts = chunks(text, 12000)
    if len(parts) <= MAX_CONDITION_CHUNKS:
        return parts
    selected: list[str] = []
    keywords = (
        "scope", "deviation", "precede", "material substitution", "branch connection", "project requirements",
        "notes:", "appendix b", "certificate", "inspection", "testing", "welding", "bolting", "gasket", "flange",
    )
    for index, part in enumerate(parts):
        folded = _fold(part)
        if index < 2 or index >= len(parts) - 2 or any(keyword in folded for keyword in keywords):
            selected.append(part)
        if len(selected) >= MAX_CONDITION_CHUNKS:
            break
    return selected


async def _extract_temporary_conditions(conditionals: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    rules: list[dict[str, Any]] = []
    unverifiable: list[dict[str, Any]] = []
    for entry in conditionals[:MAX_CONDITION_FILES]:
        path = clean(entry.get("path") or entry.get("filename") or "condicional")
        text = entry_text(entry)
        if not text:
            unverifiable.append({
                "topic": "Condicional sem texto extraível",
                "reason": "O documento não forneceu texto suficiente; revisar manualmente ou fornecer PDF pesquisável.",
                "source_document": path,
            })
            continue
        selected_parts = _condition_selection(text)
        if len(chunks(text, 12000)) > len(selected_parts):
            unverifiable.append({
                "topic": "Cobertura parcial da condicional",
                "reason": f"Documento extenso; foram priorizados {len(selected_parts)} blocos relevantes nesta execução.",
                "source_document": path,
            })
        for index, part in enumerate(selected_parts, 1):
            system = (
                "Você converte especificações industriais de um cliente em regras condicionais rastreáveis. "
                "Nunca transforme exemplo, índice, histórico de revisão ou regra de outra classe em obrigação geral. "
                "Retorne somente JSON válido."
            )
            user = f"""CLIENTE/PROJETO: documento exclusivo da execução atual
DOCUMENTO: {path}
BLOCO: {index}/{len(selected_parts)}

TEXTO
{part}

Extraia somente condicionais verificáveis. Para cada regra informe exatamente quando ela se aplica, o que exige, exceções e evidência. Não aplique a outro cliente ou projeto.
Formato:
{{"conditional_rules":[{{"category":"scope|document_control|material|piping|welding|bolting|flange|valve|quality|testing|certification|coating|other","when":{{"area":[],"class_codes":[],"service_codes":[],"components":[],"condition":""}},"then":[],"exceptions":[],"mandatory":true,"source_document":"{path}","source_location":"página/seção/bloco","source_evidence":"trecho curto"}}],"not_verifiable":[{{"topic":"","reason":"","source_document":"{path}"}}]}}"""
            try:
                result = await model(system, user, 3000)
                rules.extend(objects(result.get("conditional_rules")))
                unverifiable.extend(objects(result.get("not_verifiable")))
            except HTTPException as exc:
                unverifiable.append({
                    "topic": "Extração de condicionais",
                    "reason": str(exc.detail),
                    "source_document": path,
                })
    unique: list[dict[str, Any]] = []
    seen: set[str] = set()
    for rule in rules:
        key = "|".join([
            clean(rule.get("source_document")).casefold(),
            clean(rule.get("source_location")).casefold(),
            clean(rule.get("source_evidence")).casefold(),
            clean(rule.get("then")).casefold(),
        ])
        if key.strip("|") and key not in seen:
            seen.add(key)
            unique.append(rule)
    return [{**rule, "conditional_id": f"COND-{index:03d}"} for index, rule in enumerate(unique, 1)], unverifiable


def _client_context(
    opportunity: dict[str, Any],
    package: dict[str, Any],
    drawings: list[dict[str, Any]],
    conditionals: list[dict[str, Any]],
    temporary_rules: list[dict[str, Any]],
    profile_id: str,
) -> dict[str, Any]:
    document_summary = [
        {
            "source_document": entry.get("path"),
            "summary": entry_text(entry)[:1800],
        }
        for entry in [*drawings, *conditionals]
    ]
    extraction = {
        "document_summary": document_summary,
        "requirements": [
            {
                "requirement": rule.get("then"),
                "source_document": rule.get("source_document"),
                "source_evidence": rule.get("source_evidence"),
            }
            for rule in temporary_rules
        ],
        "commitments": [],
        "drawing_analysis": [],
    }
    enriched = dict(opportunity)
    if profile_id:
        selected = _profile_by_id(profile_id)
        if selected:
            identity = selected.get("identity") or {}
            enriched["client"] = " ".join([
                clean(enriched.get("client")),
                clean(identity.get("engineering_client")),
                clean(identity.get("operator_client")),
                clean(identity.get("project")),
                clean(identity.get("project_code")),
            ])
    context = resolve_client_knowledge(enriched, extraction)
    if not context.get("matched"):
        return context

    drawing_text = "\n".join(entry_text(entry)[:12000] for entry in drawings)
    profile = _profile_by_id(str(context.get("profile_id") or ""))
    if profile:
        area = _detect_area(enriched, drawing_text.casefold())
        class_codes = _detected_classes(drawing_text)
        service_codes = _detected_services(profile, drawing_text.casefold())
        pmcs = profile.get("pmc_profiles") if isinstance(profile.get("pmc_profiles"), dict) else {}
        context["inferred_context"] = {
            "area": area,
            "class_codes": class_codes,
            "service_codes": service_codes,
        }
        context["rules"] = _active_rules(profile, area, class_codes, service_codes)
        context["pmc_profiles"] = {
            key: value for key, value in pmcs.items()
            if any(key.casefold().endswith(f":{code.casefold()}") for code in class_codes)
        }
        context["alerts"] = _context_alerts(profile, area, class_codes, service_codes)
    return context


def _context_alert_outputs(context: dict[str, Any]) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]]]:
    findings: list[dict[str, Any]] = []
    corrections: list[dict[str, Any]] = []
    unverifiable: list[dict[str, Any]] = []
    for alert in context.get("alerts") or []:
        if alert.get("status") != "candidate_finding":
            unverifiable.append({
                "topic": alert.get("check_id") or "Regra do cliente",
                "reason": alert.get("message"),
                "source_document": context.get("profile_id"),
            })
            continue
        finding = {
            "severity": alert.get("severity") or "high",
            "category": "technical",
            "title": alert.get("check_id") or "Incompatibilidade com perfil do cliente",
            "inconsistency": alert.get("message"),
            "impact": "A classe, serviço ou área pode estar sendo interpretada sob uma regra de outro contexto.",
            "client_evidence": f"Perfil {context.get('profile_id')}",
            "step_evidence": "",
            "required_correction": "Confirmar a classe, o serviço, a área e a aprovação de engenharia antes da liberação do desenho.",
            "blocking": False,
            "requires_human_validation": True,
            "drawing_rule_id": alert.get("check_id"),
        }
        findings.append(finding)
        corrections.append({
            "section": "Base de projeto e classe de tubulação",
            "current_text": alert.get("message"),
            "corrected_text": finding["required_correction"],
            "reason": f"Regra específica do cliente {context.get('profile_id')}",
            "requires_human_validation": True,
        })
    return findings, corrections, unverifiable


def _summary(drawings: list[dict[str, Any]], findings: list[dict[str, Any]], conditionals: list[dict[str, Any]], context: dict[str, Any]) -> dict[str, Any]:
    analyzed = sum(1 for item in drawings if item.get("status") == "analyzed")
    pages = sum(int(item.get("pages_analyzed") or 0) for item in drawings)
    blockers = sum(1 for item in findings if item.get("blocking"))
    severities = {str(item.get("severity") or "").casefold() for item in findings}
    if blockers or "critical" in severities:
        risk, recommendation = "critical", "do_not_submit"
    elif "high" in severities:
        risk, recommendation = "high", "review_before_submit"
    elif "medium" in severities:
        risk, recommendation = "medium", "submit_with_reservations"
    else:
        risk, recommendation = "low", "submit"
    coverage = round(analyzed / len(drawings) * 100, 1) if drawings else 0.0
    profile_label = context.get("profile_id") or "sem perfil permanente"
    return {
        "recommendation": recommendation,
        "risk_level": risk,
        "executive_opinion": (
            f"Foram analisados {analyzed} de {len(drawings)} desenho(s), totalizando {pages} página(s). "
            f"Foram identificados {len(findings)} achado(s), sendo {blockers} bloqueante(s). "
            f"Base condicional aplicada: {profile_label}; condicionais temporárias: {len(conditionals)}."
        ),
        "adherence_percent": None,
        "coverage_percent": coverage,
        "findings_total": len(findings),
        "blocking_risks": blockers,
        "drawings_total": len(drawings),
        "drawings_analyzed": analyzed,
        "drawing_pages_analyzed": pages,
        "client_profile_id": context.get("profile_id"),
        "temporary_conditionals": len(conditionals),
    }


def _paragraph(value: Any, style: Any) -> Paragraph:
    return Paragraph(escape(clean(value)).replace("\n", "<br/>"), style)


def _write_pdf(path: Path, data: dict[str, Any]) -> None:
    styles = getSampleStyleSheet()
    small = ParagraphStyle("drawing-small", parent=styles["BodyText"], fontSize=7, leading=9)
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=10 * mm, leftMargin=10 * mm, topMargin=10 * mm, bottomMargin=10 * mm)
    summary = data["summary"]
    opportunity = data.get("opportunity") or {}
    story: list[Any] = [
        Paragraph("RELATÓRIO DE ANÁLISE DE DESENHOS", styles["Title"]),
        Spacer(1, 3 * mm),
        _paragraph(f"Referência: {opportunity.get('opportunity_id', '')} · Cliente: {opportunity.get('client', '')}", styles["Heading3"]),
        _paragraph(f"Perfil condicional: {summary.get('client_profile_id') or 'não identificado'}", styles["Heading3"]),
        _paragraph(summary.get("executive_opinion"), styles["BodyText"]),
        Spacer(1, 4 * mm),
    ]
    rows: list[list[Any]] = [["Severidade", "Desenho / local", "Achado", "Evidência", "Correção"]]
    for item in data.get("findings") or []:
        rows.append([
            _paragraph(item.get("severity"), small),
            _paragraph(f"{item.get('source_document', '')} {item.get('source_location', '')}", small),
            _paragraph(item.get("title") or item.get("inconsistency"), small),
            _paragraph(item.get("client_evidence") or item.get("step_evidence"), small),
            _paragraph(item.get("required_correction"), small),
        ])
    if len(rows) == 1:
        rows.append(["—", "—", "Nenhum achado confirmado", "—", "Manter revisão humana final"])
    table = Table(rows, repeatRows=1, colWidths=[24 * mm, 48 * mm, 62 * mm, 74 * mm, 68 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0B2D4D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(table)
    story.append(PageBreak())
    story.append(Paragraph("Condicionais aplicadas", styles["Heading1"]))
    for rule in data.get("temporary_conditions") or []:
        story.append(_paragraph(f"{rule.get('conditional_id')} · {rule.get('source_document')} · quando {rule.get('when')} → {rule.get('then')}", small))
    if not data.get("temporary_conditions"):
        story.append(_paragraph("Nenhuma condicional temporária foi extraída do ZIP.", styles["BodyText"]))
    doc.build(story)


def _write_xlsx(path: Path, data: dict[str, Any]) -> None:
    workbook = Workbook()
    navy = PatternFill("solid", fgColor="0B2D4D")
    white = Font(color="FFFFFF", bold=True)

    def add_sheet(name: str, headers: list[str], rows: list[list[Any]], widths: list[int]) -> None:
        sheet = workbook.active if len(workbook.sheetnames) == 1 and workbook.active.title == "Sheet" else workbook.create_sheet()
        sheet.title = name
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = navy
            cell.font = white
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in rows:
            sheet.append([clean(value) for value in row])
        for index, width in enumerate(widths, 1):
            sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

    summary = data["summary"]
    add_sheet("Resumo", ["Indicador", "Valor"], [[key, value] for key, value in summary.items()], [38, 110])
    add_sheet(
        "Achados",
        ["ID", "Severidade", "Regra", "Documento", "Localização", "Achado", "Evidência", "Correção", "Bloqueante", "Validação humana"],
        [[
            item.get("id"), item.get("severity"), item.get("drawing_rule_id"), item.get("source_document"), item.get("source_location"),
            item.get("title") or item.get("inconsistency"), item.get("client_evidence") or item.get("step_evidence"),
            item.get("required_correction"), "SIM" if item.get("blocking") else "NÃO", "SIM" if item.get("requires_human_validation") else "NÃO",
        ] for item in data.get("findings") or []],
        [12, 14, 18, 42, 24, 50, 65, 65, 14, 18],
    )
    add_sheet(
        "Condicionais",
        ["ID", "Categoria", "Quando se aplica", "Exigência", "Exceções", "Documento", "Localização", "Evidência"],
        [[
            item.get("conditional_id"), item.get("category"), item.get("when"), item.get("then"), item.get("exceptions"),
            item.get("source_document"), item.get("source_location"), item.get("source_evidence"),
        ] for item in data.get("temporary_conditions") or []],
        [14, 20, 58, 62, 45, 42, 28, 65],
    )
    add_sheet(
        "Não verificáveis",
        ["Tema", "Motivo", "Documento", "Localização", "Evidência"],
        [[item.get("topic"), item.get("reason"), item.get("source_document"), item.get("source_location"), item.get("source_evidence")] for item in data.get("not_verifiable") or []],
        [34, 75, 42, 28, 60],
    )
    workbook.save(path)


def _generate_artifacts(opportunity: dict[str, Any], package: dict[str, Any], analysis: dict[str, Any]) -> dict[str, Any]:
    oid = clean(opportunity.get("opportunity_id")) or safe(Path(str(package.get("package_name") or "drawing-audit")).stem, "drawing-audit")
    root = workspace(oid)
    data = {
        "opportunity": opportunity,
        "package": {
            "package_name": package.get("package_name"),
            "summary": package.get("summary") or {},
        },
        **analysis,
        "generated_at": datetime.now(UTC).isoformat(),
    }
    json_path = root / safe(f"Analise_Desenhos_{oid}.json")
    xlsx_path = root / safe(f"Analise_Desenhos_{oid}.xlsx")
    pdf_path = root / safe(f"Relatorio_Desenhos_{oid}.pdf")
    json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    _write_xlsx(xlsx_path, data)
    _write_pdf(pdf_path, data)
    artifacts = [artifact(oid, path) for path in (pdf_path, xlsx_path, json_path)]
    return {
        "status": "analysis_completed",
        "analysis_type": "drawing_audit",
        "opportunity": opportunity,
        "summary": data["summary"],
        "findings": data["findings"],
        "corrections": data["corrections"],
        "requirements": [],
        "commitments": [],
        "not_verifiable": data["not_verifiable"],
        "client_knowledge": data["client_knowledge"],
        "temporary_conditions": data["temporary_conditions"],
        "drawing_analysis": data["drawing_analysis"],
        "artifacts": artifacts,
        "completed_at": datetime.now(UTC).isoformat(),
    }


@app.post("/v1/drawings/from-package", dependencies=[Depends(auth)])
async def drawings_from_package(
    file: UploadFile = File(...),
    opportunity_id: str = Form(default=""),
    client: str = Form(default=""),
    rfq_id: str = Form(default=""),
    owner: str = Form(default=""),
    agents_json: str = Form(default="[]"),
) -> dict[str, Any]:
    if not (file.filename or "").casefold().endswith(".zip"):
        raise HTTPException(status_code=415, detail="Envie um único arquivo ZIP")
    try:
        agents = json.loads(agents_json)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=422, detail="agents_json inválido") from exc
    if not isinstance(agents, list) or "drawing" not in agents:
        raise HTTPException(status_code=422, detail="Modo drawing não informado")

    raw = await file.read()
    package = analyze_package(file.filename or "drawings.zip", raw, opportunity_id.strip() or None, include_content=True)
    drawings, conditionals, other = _classify_drawing_package(package)
    if not drawings:
        raise HTTPException(status_code=422, detail="Nenhum PDF ou imagem de desenho foi identificado no ZIP")

    metadata = _agent_metadata(agents)
    opportunity = {
        "opportunity_id": opportunity_id.strip() or package.get("opportunity_id"),
        "client": client.strip() or (package.get("inferred") or {}).get("client") or "CLIENTE NÃO IDENTIFICADO",
        "rfq_id": rfq_id.strip(),
        "owner": owner.strip(),
        "agents": agents,
        "area": metadata.get("area") or "",
        "project": metadata.get("project") or "",
        "client_profile_id": metadata.get("profile_id") or "",
    }

    preparation_warnings = _prepare_missing_visuals(raw, drawings, str(opportunity["opportunity_id"]))
    temporary_rules, conditional_unverifiable = await _extract_temporary_conditions(conditionals)
    context = _client_context(
        opportunity,
        package,
        drawings[:MAX_DRAWINGS],
        conditionals,
        temporary_rules,
        metadata.get("profile_id") or "",
    )

    drawing_analysis: list[dict[str, Any]] = []
    for entry in drawings[:MAX_DRAWINGS]:
        drawing_analysis.append(await analyze_drawing_with_client_context(entry, context, temporary_rules))

    findings, corrections, visual_unverifiable = drawing_findings(drawing_analysis)
    context_findings, context_corrections, context_unverifiable = _context_alert_outputs(context)
    findings.extend(context_findings)
    corrections.extend(context_corrections)
    for index, item in enumerate(findings, 1):
        item["id"] = f"D-F-{index:03d}"
    for index, item in enumerate(corrections, 1):
        item["id"] = f"D-C-{index:03d}"

    non_visual = [
        {
            "topic": "Arquivo não classificado na análise de desenhos",
            "reason": "O arquivo não foi identificado como desenho nem como condicional do cliente.",
            "source_document": entry.get("path"),
        }
        for entry in other
    ]
    warnings = [
        {"topic": "Preparação de desenhos", "reason": warning, "source_document": package.get("package_name")}
        for warning in preparation_warnings
    ]
    if not context.get("matched"):
        warnings.append({
            "topic": "Perfil permanente do cliente",
            "reason": "Nenhum perfil permanente foi reconhecido. Foram aplicadas somente as condicionais do ZIP e as regras gerais de desenho.",
            "source_document": package.get("package_name"),
        })

    not_verifiable = [
        *conditional_unverifiable,
        *visual_unverifiable,
        *context_unverifiable,
        *non_visual,
        *warnings,
    ]
    analysis = {
        "summary": _summary(drawing_analysis, findings, temporary_rules, context),
        "findings": findings,
        "corrections": corrections,
        "not_verifiable": not_verifiable,
        "client_knowledge": {
            "matched": bool(context.get("matched")),
            "profile_id": context.get("profile_id"),
            "identity": context.get("identity") or {},
            "inferred_context": context.get("inferred_context") or {},
            "rules_loaded": len(context.get("rules") or []),
            "pmc_profiles_loaded": sorted((context.get("pmc_profiles") or {}).keys()),
            "source_registry": context.get("source_registry") or [],
        },
        "temporary_conditions": temporary_rules,
        "drawing_analysis": drawing_analysis,
    }
    return _generate_artifacts(opportunity, package, analysis)
